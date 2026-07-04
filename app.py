from flask import Flask, render_template, request, redirect, url_for, session, Response
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from urllib.parse import urlencode, quote_plus

app = Flask(__name__)

# Required for login sessions
app.secret_key = "temporary-dev-secret-key-change-later"


# =============================
# TEMPORARY GUEST DATA
# Later this can move to a database
# =============================

GUESTS = {
    "prithvi7007@gmail.com": {
        "password": "123",
        "name": "Prithvi",
        "allowed_events": [
            "haldi",
            "hindu_wedding",
            "christian_wedding_reception"
        ],
        "max_guests": 2
    },

    "family@example.com": {
        "password": "123",
        "name": "Family Guest",
        "allowed_events": [
            "haldi",
            "hindu_wedding",
            "christian_wedding_reception"
        ],
        "max_guests": 4
    },

    "friends@example.com": {
        "password": "123",
        "name": "Friend Guest",
        "allowed_events": [
            "christian_wedding_reception"
        ],
        "max_guests": 1
    }
}


# =============================
# EVENT DATA
# Later this can move to a database
# =============================

EVENTS = {
    "haldi": {
        "id": "haldi",
        "title": "Haldi",
        "date": "November 19, 2026",
        "time": "11:00 AM",
        "calendar_start": "20261119T110000",
        "calendar_end": "20261119T140000",
        "short_date": "Nov 19",
        "badge_title": "Haldi",
        "venue_name": "Private Residence",
        "location": "5312 Windingbrook Trail, Wesley Chapel, FL 33544",
        "description": "Join us for a joyful Haldi celebration with family, friends, traditions, and blessings.",
        "timeline": [
            "11:00 AM - Haldi celebration begins",
            "2:00 PM - Celebration concludes"
        ],
        "attire_image": "images/schedule/attire-haldi.png",
        "attire_heading": "Pastel & Floral Shades",
        "attire_subheading": "Daytime Event",
        "attire": "Festive Indian attire encouraged. Bright yellows, oranges, greens, and floral colors are perfect for the Haldi celebration."
    },


    "hindu_wedding": {
        "id": "hindu_wedding",
        "title": "Traditional Telugu Wedding",
        "date": "November 20, 2026",
        "time": "9:00 AM",
        "calendar_start": "20261120T090000",
        "calendar_end": "20261120T120000",
        "short_date": "Nov 20",
        "badge_title": "Hindu Ceremony",
        "venue_name": "Hindu Wedding Venue",
        "location": "9338 Old Gibsonton Dr, Gibsonton, FL 33534",
        "description": "Join us for the Hindu wedding ceremony as we celebrate our marriage with traditional rituals and blessings.",
        "timeline": [
            "9:00 AM - Ceremony begins",
            "12:00 PM - Ceremony concludes"
        ],
        "attire_image": "images/schedule/attire-hindu.png",
        "attire_heading": "Traditional Formal Wear",
        "attire_subheading": "Ceremony Event",
        "attire": "Traditional Indian formal attire is encouraged. Sarees, lehengas, anarkalis, kurtas, sherwanis, or formal festive wear are welcome."
    },

    "christian_wedding_reception": {
        "id": "christian_wedding_reception",
        "title": "Church Wedding and Reception",
        "date": "November 21, 2026",
        "time": "5:00 PM",
        "calendar_start": "20261121T170000",
        "calendar_end": "20261121T230000",
        "short_date": "Nov 21",
        "badge_title": "Reception",
        "venue_name": "Church Wedding and Reception Venue",
        "location": "9724 Cross Creek Blvd, Tampa, FL 33647",
        "description": "Join us for the Christian wedding ceremony followed by an evening reception, dinner, dancing, and celebration.",
        "timeline": [
            "5:00 PM - Ceremony begins",
            "6:00 PM - Reception begins",
            "11:00 PM - Celebration concludes"
        ],
        "attire_image": "images/schedule/attire-reception.png",
        "attire_heading": "Formal Evening Attire",
        "attire_subheading": "Ceremony & Reception",
        "attire": "Formal or semi-formal wedding attire. Evening dresses, suits, dress shirts, cocktail attire, and elegant traditional outfits are welcome."
    }
}


# =============================
# EXCEL RSVP SETUP
# =============================

DATA_DIR = Path("data")
RSVP_FILE = DATA_DIR / "rsvps.xlsx"

RSVP_HEADERS = [
    "updated_at",
    "guest_email",
    "guest_name",
    "event_id",
    "event_title",
    "attending",
    "guest_count",
    "notes"
]


def ensure_rsvp_workbook():
    DATA_DIR.mkdir(exist_ok=True)

    if not RSVP_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "RSVPs"
        ws.append(RSVP_HEADERS)
        wb.save(RSVP_FILE)


def save_rsvp_to_excel(guest_email, guest_name, event_id, event_title, attending, guest_count, notes):
    ensure_rsvp_workbook()

    wb = load_workbook(RSVP_FILE)
    ws = wb["RSVPs"]

    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # If this guest already RSVP'd to this event, update the row instead of creating duplicates
    existing_row = None

    for row_num in range(2, ws.max_row + 1):
        existing_email = ws.cell(row=row_num, column=2).value
        existing_event_id = ws.cell(row=row_num, column=4).value

        if existing_email == guest_email and existing_event_id == event_id:
            existing_row = row_num
            break

    row_values = [
        updated_at,
        guest_email,
        guest_name,
        event_id,
        event_title,
        attending,
        guest_count,
        notes
    ]

    if existing_row:
        for col_num, value in enumerate(row_values, start=1):
            ws.cell(row=existing_row, column=col_num).value = value
    else:
        ws.append(row_values)

    wb.save(RSVP_FILE)


def get_saved_rsvps_for_guest(guest_email):
    ensure_rsvp_workbook()

    wb = load_workbook(RSVP_FILE)
    ws = wb["RSVPs"]

    saved = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        updated_at, email, guest_name, event_id, event_title, attending, guest_count, notes = row

        if email == guest_email:
            saved[event_id] = {
                "attending": attending,
                "guest_count": guest_count,
                "notes": notes
            }

    return saved


def get_current_guest():
    guest_email = session.get("guest_email")

    if not guest_email:
        return None

    guest = GUESTS.get(guest_email)

    if not guest:
        return None

    guest_data = guest.copy()
    guest_data["email"] = guest_email

    return guest_data

def build_google_calendar_link(event):
    calendar_params = {
        "action": "TEMPLATE",
        "text": event["title"],
        "dates": f"{event['calendar_start']}/{event['calendar_end']}",
        "details": event["description"],
        "location": event["location"],
        "ctz": "America/New_York"
    }

    return "https://calendar.google.com/calendar/render?" + urlencode(calendar_params)


def build_directions_link(event):
    return "https://www.google.com/maps/search/?api=1&query=" + quote_plus(event["location"])

def escape_ics_text(text):
    return (
        text.replace("\\", "\\\\")
            .replace(",", "\\,")
            .replace(";", "\\;")
            .replace("\n", "\\n")
    )


# =============================
# ROUTES
# =============================

@app.route("/")
def login_page():
    return render_template("index.html")


@app.route("/login", methods=["POST"])
def login():
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")

    guest = GUESTS.get(email)

    if guest and guest["password"] == password:
        session["guest_email"] = email
        session["guest_name"] = guest["name"]
        return redirect(url_for("dashboard", login="success"))

    return redirect(url_for("login_page"))


@app.route("/dashboard")
def dashboard():
    guest = get_current_guest()

    login_effect = request.args.get("login")

    if not guest:
        return redirect(url_for("login_page"))

    allowed_events = []

    for event_id in guest["allowed_events"]:
        if event_id in EVENTS:
            event = EVENTS[event_id].copy()
            event["calendar_link"] = build_google_calendar_link(event)
            event["directions_link"] = build_directions_link(event)
            allowed_events.append(event)
    saved_rsvps = get_saved_rsvps_for_guest(guest["email"])

    celebrate = request.args.get("celebrate")
    celebrate_event_id = request.args.get("event")

    return render_template(
        "dashboard.html",
        guest=guest,
        events=allowed_events,
        saved_rsvps=saved_rsvps,
        celebrate=celebrate,
        celebrate_event_id=celebrate_event_id,
        login_effect=login_effect
    )

@app.route("/rsvp", methods=["POST"])
def save_rsvp():
    guest = get_current_guest()

    if not guest:
        return redirect(url_for("login_page"))

    event_id = request.form.get("event_id")

    # Security check: guest can only RSVP to events they are invited to
    if event_id not in guest["allowed_events"]:
        return redirect(url_for("dashboard"))

    event = EVENTS.get(event_id)

    if not event:
        return redirect(url_for("dashboard"))

    attending = request.form.get("attending", "")
    guest_count = request.form.get("guest_count", "0")
    notes = request.form.get("notes", "").strip()

    try:
        guest_count = int(guest_count)
    except ValueError:
        guest_count = 0
    
    if attending == "No":
        guest_count = 0

    guest_count = max(0, min(guest_count, guest["max_guests"]))

    save_rsvp_to_excel(
        guest_email=guest["email"],
        guest_name=guest["name"],
        event_id=event["id"],
        event_title=event["title"],
        attending=attending,
        guest_count=guest_count,
        notes=notes
    )

    if attending == "Yes":
        return redirect(url_for("dashboard", celebrate="yes", event=event["id"]) + "#schedule")

    return redirect(url_for("dashboard") + "#schedule")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

@app.route("/calendar/<event_id>.ics")
def apple_calendar_event(event_id):
    guest = get_current_guest()

    if not guest:
        return redirect(url_for("login_page"))

    # Security check: guest can only download calendar files for invited events
    if event_id not in guest["allowed_events"]:
        return redirect(url_for("dashboard"))

    event = EVENTS.get(event_id)

    if not event:
        return redirect(url_for("dashboard"))

    title = escape_ics_text(event["title"])
    description = escape_ics_text(event["description"])
    location = escape_ics_text(event["location"])

    ics_content = "\r\n".join([
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Adlin and Prithvi Wedding//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "BEGIN:VEVENT",
        f"UID:{event_id}@adlin-prithvi-wedding",
        "DTSTAMP:20260701T120000Z",
        f"DTSTART;TZID=America/New_York:{event['calendar_start']}",
        f"DTEND;TZID=America/New_York:{event['calendar_end']}",
        f"SUMMARY:{title}",
        f"DESCRIPTION:{description}",
        f"LOCATION:{location}",
        "END:VEVENT",
        "END:VCALENDAR",
        ""
    ])

    filename = f"{event_id}.ics"

    return Response(
        ics_content,
        mimetype="text/calendar",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


if __name__ == "__main__":
    app.run(debug=True)
    
