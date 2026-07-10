import csv
import hashlib
import os
import re
import secrets
import sys
from pathlib import Path

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
load_dotenv(PROJECT_ROOT / ".env")

BASE_INVITE_URL = os.getenv("BASE_INVITE_URL", "https://adlinprithvi.cloud/invite")

EVENT_MAP = {
    "Haldi_Count": "haldi",
    "Hindu_Count": "telugu_wedding",
    "Church_Count": "church_wedding",
}


def get_db_connection():
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "127.0.0.1"),
        port=os.getenv("DB_PORT", "5433"),
        dbname=os.getenv("DB_NAME", "wedding_db"),
        user=os.getenv("DB_USER", "wedding_user"),
        password=os.getenv("DB_PASSWORD"),
    )


def clean_text(value):
    if value is None:
        return ""

    value = str(value).strip()

    if value.lower() in {"none", "nan", "null"}:
        return ""

    return value


def clean_count(value):
    if value is None or value == "":
        return 0

    try:
        return int(float(value))
    except ValueError:
        return 0


def slugify(value):
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value)
    return value.strip("-") or "guest"


def make_source_key(row):
    raw = "|".join(
        [
            row["Represent"],
            row["First Name"],
            row["Last Name"],
            row["Partner Name"],
            row["Display Name"],
            row["GROUP"],
        ]
    )

    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def make_invite_token(display_name, first_name, last_name):
    base = display_name or f"{first_name} {last_name}".strip() or "guest"
    token_base = slugify(base)
    suffix = secrets.token_hex(3)
    return f"{token_base}-{suffix}"


def read_excel_rows(excel_path):
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    headers = []
    for cell in sheet[1]:
        headers.append(clean_text(cell.value))

    required_columns = [
        "Represent",
        "First Name",
        "Last Name",
        "Partner Name",
        "Display Name",
        "GROUP",
        "Message",
        "Haldi_Count",
        "Hindu_Count",
        "Church_Count",
    ]

    missing = [col for col in required_columns if col not in headers]

    if missing:
        raise ValueError(f"Missing columns in Excel: {missing}")

    rows = []

    for row_values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, row_values))

        cleaned = {
            "Represent": clean_text(row.get("Represent")),
            "First Name": clean_text(row.get("First Name")),
            "Last Name": clean_text(row.get("Last Name")),
            "Partner Name": clean_text(row.get("Partner Name")),
            "Display Name": clean_text(row.get("Display Name")),
            "GROUP": clean_text(row.get("GROUP")),
            "Message": clean_text(row.get("Message")),
            "Haldi_Count": clean_count(row.get("Haldi_Count")),
            "Hindu_Count": clean_count(row.get("Hindu_Count")),
            "Church_Count": clean_count(row.get("Church_Count")),
        }

        # Skip fully blank rows
        if not any(
            [
                cleaned["First Name"],
                cleaned["Last Name"],
                cleaned["Partner Name"],
                cleaned["Display Name"],
            ]
        ):
            continue

        if not cleaned["Display Name"]:
            names = [cleaned["First Name"], cleaned["Partner Name"]]
            cleaned["Display Name"] = " & ".join([name for name in names if name])

        cleaned["source_key"] = make_source_key(cleaned)

        rows.append(cleaned)

    return rows


def verify_events_exist(cur):
    expected_event_ids = list(EVENT_MAP.values())

    cur.execute(
        """
        SELECT event_id
        FROM events
        WHERE event_id = ANY(%s)
        """,
        (expected_event_ids,),
    )

    found = {row[0] for row in cur.fetchall()}
    missing = set(expected_event_ids) - found

    if missing:
        raise ValueError(f"These event_id values are missing from events table: {sorted(missing)}")


def import_guests(excel_path):
    rows = read_excel_rows(excel_path)

    if not rows:
        print("No guest rows found.")
        return

    output_path = PROJECT_ROOT / "data" / "generated_invite_links.csv"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    conn = get_db_connection()

    imported = []

    try:
        with conn:
            with conn.cursor() as cur:
                verify_events_exist(cur)

                for row in rows:
                    source_key = row["source_key"]

                    # Reuse existing invite_token if this guest was imported before
                    cur.execute(
                        """
                        SELECT invite_token
                        FROM invitations
                        WHERE source_key = %s
                        """,
                        (source_key,),
                    )

                    existing = cur.fetchone()

                    if existing:
                        invite_token = existing[0]
                    else:
                        invite_token = make_invite_token(
                            row["Display Name"],
                            row["First Name"],
                            row["Last Name"],
                        )

                    cur.execute(
                        """
                        INSERT INTO invitations (
                            source_key,
                            represent_side,
                            first_name,
                            last_name,
                            partner_name,
                            display_name,
                            guest_group,
                            invite_message,
                            invite_token,
                            is_active
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE)
                        ON CONFLICT (source_key)
                        DO UPDATE SET
                            represent_side = EXCLUDED.represent_side,
                            first_name = EXCLUDED.first_name,
                            last_name = EXCLUDED.last_name,
                            partner_name = EXCLUDED.partner_name,
                            display_name = EXCLUDED.display_name,
                            guest_group = EXCLUDED.guest_group,
                            invite_message = EXCLUDED.invite_message,
                            is_active = TRUE
                        RETURNING invitation_id, invite_token
                        """,
                        (
                            source_key,
                            row["Represent"],
                            row["First Name"],
                            row["Last Name"],
                            row["Partner Name"],
                            row["Display Name"],
                            row["GROUP"],
                            row["Message"],
                            invite_token,
                        ),
                    )

                    invitation_id, final_token = cur.fetchone()

                    # Rebuild event permissions from Excel counts
                    cur.execute(
                        """
                        DELETE FROM invitation_event_permissions
                        WHERE invitation_id = %s
                        """,
                        (invitation_id,),
                    )

                    event_counts = {}

                    for excel_col, event_id in EVENT_MAP.items():
                        max_guests = row[excel_col]
                        event_counts[event_id] = max_guests

                        if max_guests > 0:
                            cur.execute(
                                """
                                INSERT INTO invitation_event_permissions (
                                    invitation_id,
                                    event_id,
                                    max_guests
                                )
                                VALUES (%s, %s, %s)
                                ON CONFLICT (invitation_id, event_id)
                                DO UPDATE SET max_guests = EXCLUDED.max_guests
                                """,
                                (invitation_id, event_id, max_guests),
                            )

                    invite_url = f"{BASE_INVITE_URL}/{final_token}"

                    imported.append(
                        {
                            "Display Name": row["Display Name"],
                            "First Name": row["First Name"],
                            "Last Name": row["Last Name"],
                            "Partner Name": row["Partner Name"],
                            "GROUP": row["GROUP"],
                            "Haldi_Count": event_counts["haldi"],
                            "Hindu_Count": event_counts["telugu_wedding"],
                            "Church_Count": event_counts["church_wedding"],
                            "Invite URL": invite_url,
                        }
                    )

        with output_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "Display Name",
                    "First Name",
                    "Last Name",
                    "Partner Name",
                    "GROUP",
                    "Haldi_Count",
                    "Hindu_Count",
                    "Church_Count",
                    "Invite URL",
                ],
            )
            writer.writeheader()
            writer.writerows(imported)

        print(f"Imported invitations: {len(imported)}")
        print(f"Invite links saved to: {output_path}")

    finally:
        conn.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_guests.py data/guest_list.xlsx")
        sys.exit(1)

    excel_path = PROJECT_ROOT / sys.argv[1]

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    import_guests(excel_path)